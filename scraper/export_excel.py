"""Export the latest snapshot data to a clean multi-sheet Excel viewer.

Sheets:
- ALL — every tracked product, sorted by set + availability + price
- AVAILABLE NOW — only in-stock products, the actionable buy list
- CHANGES — restocks, stockouts, price moves since the previous snapshot
- CHEAPEST PER SET — the single cheapest in-stock listing per set
- SUMMARY — counts per set / platform / shop

Color coding on the FLAG column:
  RESTOCK = green, STOCKOUT = red, PRICE↓ = light green, PRICE↑ = light red.
"""
import html
import pandas as pd
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.formatting.rule import CellIsRule
from openpyxl.utils import get_column_letter
from .db import connect
from .config import DATA_DIR


def _clean_title(t):
    if not isinstance(t, str): return t
    return html.unescape(t).replace("–", "-").replace("—", "-").strip()


def load_latest():
    """Latest snapshot per product + previous snapshot for change detection."""
    with connect() as conn:
        df = pd.read_sql_query("""
            WITH ranked AS (
                SELECT s.product_id, s.observed_at, s.price_eur, s.available,
                       p.platform, p.set_code, p.game, p.shop, p.title, p.url,
                       p.first_seen_at,
                       ROW_NUMBER() OVER (PARTITION BY s.product_id
                                          ORDER BY s.observed_at DESC) rn
                FROM snapshots s
                JOIN products p ON p.id = s.product_id
            ),
            latest AS (SELECT * FROM ranked WHERE rn = 1),
            prev   AS (SELECT product_id, price_eur AS price_prev,
                              available AS avail_prev
                       FROM ranked WHERE rn = 2)
            SELECT
                l.product_id, l.platform, l.set_code, l.game, l.shop, l.title, l.url,
                l.price_eur AS price_now,
                l.available AS avail_now,
                p.price_prev, p.avail_prev,
                l.observed_at AS last_observed,
                l.first_seen_at
            FROM latest l
            LEFT JOIN prev p USING (product_id)
        """, conn)
    df["title"] = df["title"].apply(_clean_title)
    return df


def classify_kind(price, title):
    """Returns 'case' | 'display' | 'booster' | 'accessory' | 'unknown'.
    A *case* is a carton of multiple displays (typically >=6, price >=1000€).
    'Case de 24 boosters sous blisters' (~150€) is a display variant, NOT a case."""
    import re as _re
    t = (title or "").lower()
    ACCESSORY_KW = ["protection", "plexi", "écrin", "ecrin", "support pour",
                    "rangement", "présentoir", "presentoir"]
    if any(k in t for k in ACCESSORY_KW):
        return "accessory"
    has_case_word = bool(_re.search(r"\bcase\b", t)) or "carton" in t
    if price is not None and price >= 1000:
        return "case"
    if has_case_word and "display" in t:
        return "case"
    if has_case_word and ("booster" in t or "blister" in t):
        return "display"  # case-of-boosters → display variant
    DISPLAY_KW = ["display", "boîte de booster", "boite de booster",
                  "booster box", "boosterbox", "boîte 24", "boite 24",
                  "boîte de 24", "boite de 24", "boîte de 20", "boite de 20"]
    if any(k in t for k in DISPLAY_KW):
        return "display"
    if "booster" in t and "box" not in t:
        return "booster"
    if price is None: return "unknown"
    if price <= 15: return "booster"
    if price >= 100: return "display"
    return "unknown"


def add_flag_and_status(df):
    def flag(r):
        an, ap = r["avail_now"], r["avail_prev"]
        if pd.notna(ap) and an == 1 and ap == 0: return "RESTOCK"
        if pd.notna(ap) and an == 0 and ap == 1: return "STOCKOUT"
        if pd.notna(r["price_prev"]) and pd.notna(r["price_now"]):
            d = r["price_now"] - r["price_prev"]
            if d < -0.5: return "PRICE↓"
            if d > 0.5:  return "PRICE↑"
        return ""

    def status(av):
        if av == 1: return "In Stock"
        if av == 0: return "Out"
        return "Unknown"

    df = df.copy()
    df["flag"] = df.apply(flag, axis=1)
    df["status"] = df["avail_now"].apply(status)
    df["kind"] = df.apply(lambda r: classify_kind(r["price_now"], r["title"]), axis=1)
    df["price_change"] = df.apply(
        lambda r: round(r["price_now"] - r["price_prev"], 2)
        if pd.notna(r["price_prev"]) and pd.notna(r["price_now"]) else None,
        axis=1,
    )
    return df


def order_cols(df):
    cols = ["set_code", "kind", "game", "status", "flag", "price_now", "price_change",
            "shop", "title", "platform", "last_observed", "first_seen_at", "url"]
    cols = [c for c in cols if c in df.columns]
    return df[cols].rename(columns={
        "set_code": "set", "price_now": "price_€", "price_change": "Δprice_€",
        "first_seen_at": "first_seen", "last_observed": "last_observed",
    })


def sort_avail_then_price(df):
    """Sort: set > kind (display first) > availability > price."""
    avail_rank = {"In Stock": 0, "Unknown": 1, "Out": 2}
    kind_rank  = {"display": 0, "booster": 1, "case": 2, "accessory": 3, "unknown": 4}
    df = df.copy()
    df["_r"] = df["status"].map(avail_rank)
    df["_k"] = df["kind"].map(kind_rank).fillna(99)
    df = df.sort_values(["set", "_k", "_r", "price_€"], ascending=[True, True, True, True])
    return df.drop(columns=["_r", "_k"])


def write_excel(out_path, df):
    df = add_flag_and_status(df)

    # Sheet 1: ALL — per-set sorted
    sheet_all = sort_avail_then_price(order_cols(df))

    # Sheet 2: AVAILABLE NOW — in-stock only
    sheet_avail = sort_avail_then_price(order_cols(df[df["avail_now"] == 1]))

    # Sheet 3: CHANGES — only flagged rows
    changes = df[df.apply(
        lambda r: (pd.notna(r["avail_prev"]) and r["avail_now"] != r["avail_prev"]) or
                  (pd.notna(r["price_prev"]) and pd.notna(r["price_now"])
                   and abs(r["price_now"] - r["price_prev"]) > 0.5), axis=1
    )]
    sheet_changes = order_cols(changes).sort_values(["flag", "set", "price_€"])

    # Sheet 4: CHEAPEST PER SET — one row per (set, kind) showing best in-stock deal
    # Accessories are excluded — they're not the product itself.
    in_stock = df[(df["avail_now"] == 1) & (df["kind"] != "accessory")].sort_values("price_now")
    cheapest = in_stock.groupby(["set_code", "kind"], as_index=False).first()
    sheet_cheapest = order_cols(cheapest).sort_values(["set", "kind"])

    # Sheet 5: SUMMARY — split by (set, kind) for meaningful medians
    summary_rows = []
    for (setc, kind), g in df.groupby(["set_code", "kind"]):
        in_st = (g["avail_now"] == 1).sum()
        prices_in = g.loc[g["avail_now"] == 1, "price_now"]
        summary_rows.append({
            "set": setc,
            "kind": kind,
            "total_listings": len(g),
            "in_stock": int(in_st),
            "out_of_stock": int((g["avail_now"] == 0).sum()),
            "unknown": int(g["avail_now"].isna().sum()),
            "cheapest_in_stock_€": round(prices_in.min(), 2) if len(prices_in) else None,
            "median_in_stock_€":  round(prices_in.median(), 2) if len(prices_in) else None,
        })
    summary = pd.DataFrame(summary_rows).sort_values(["set", "kind"])

    with pd.ExcelWriter(out_path, engine="openpyxl") as xw:
        sheet_all.to_excel(xw,      sheet_name="ALL",            index=False)
        sheet_avail.to_excel(xw,    sheet_name="AVAILABLE NOW",  index=False)
        sheet_changes.to_excel(xw,  sheet_name="CHANGES",        index=False)
        sheet_cheapest.to_excel(xw, sheet_name="CHEAPEST PER SET", index=False)
        summary.to_excel(xw,        sheet_name="SUMMARY",        index=False)

    # Post-format with openpyxl
    from openpyxl import load_workbook
    wb = load_workbook(out_path)

    fills = {
        "RESTOCK":  PatternFill("solid", fgColor="63BE7B"),
        "STOCKOUT": PatternFill("solid", fgColor="F8696B"),
        "PRICE↓":   PatternFill("solid", fgColor="C6EFCE"),
        "PRICE↑":   PatternFill("solid", fgColor="FFC7CE"),
    }
    status_fills = {
        "In Stock": PatternFill("solid", fgColor="E2EFDA"),
        "Out":      PatternFill("solid", fgColor="FCE4D6"),
        "Unknown":  PatternFill("solid", fgColor="F2F2F2"),
    }
    bold = Font(bold=True)

    for sheet_name in ["ALL", "AVAILABLE NOW", "CHANGES", "CHEAPEST PER SET", "SUMMARY"]:
        ws = wb[sheet_name]
        # Header row: bold + frozen
        for c in ws[1]:
            c.font = bold
            c.alignment = Alignment(horizontal="left")
        ws.freeze_panes = "A2"
        # Column widths
        widths = {"set": 12, "kind": 10, "game": 14, "status": 10, "flag": 10,
                  "price_€": 10, "Δprice_€": 10, "shop": 26, "title": 60,
                  "platform": 12, "url": 50, "last_observed": 20,
                  "first_seen": 12, "total_listings": 14, "in_stock": 10,
                  "out_of_stock": 13, "unknown": 9,
                  "cheapest_in_stock_€": 18, "median_in_stock_€": 18}
        for i, col in enumerate(ws[1], start=1):
            w = widths.get(col.value, 14)
            ws.column_dimensions[get_column_letter(i)].width = w
        # Color flag column if present
        header = [c.value for c in ws[1]]
        if "flag" in header:
            j = header.index("flag") + 1
            for row in ws.iter_rows(min_row=2, min_col=j, max_col=j):
                cell = row[0]
                if cell.value in fills:
                    cell.fill = fills[cell.value]
                    cell.font = Font(bold=True)
        if "status" in header:
            j = header.index("status") + 1
            for row in ws.iter_rows(min_row=2, min_col=j, max_col=j):
                cell = row[0]
                if cell.value in status_fills:
                    cell.fill = status_fills[cell.value]

    wb.save(out_path)


def main():
    out = DATA_DIR / "products_view.xlsx"
    df = load_latest()
    print(f"Loaded {len(df)} products from DB")
    write_excel(out, df)
    print(f"Written to: {out}")
    # Quick summary
    df = add_flag_and_status(df)
    print(f"\nFlags this run:")
    print(df["flag"].value_counts().to_string() or "  (none)")
    print(f"\nIn stock: {(df['avail_now']==1).sum()}  |  Out: {(df['avail_now']==0).sum()}  |  Unknown: {df['avail_now'].isna().sum()}")


if __name__ == "__main__":
    main()
