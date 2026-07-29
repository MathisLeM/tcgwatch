"""Hand-editable Pokemon product catalog (the "collection products" layer).

The catalog is the set of sealed SKUs we track = (set x kind x language). This
tool round-trips it through a clear Excel grid so it can be curated by hand:

  export : write data/pokemon_catalog.xlsx — one sheet per language, one row per
           set, one column per product kind. Put any mark (x / 1) in a kind cell
           to track that SKU; clear it to drop it. A sensible baseline is
           pre-filled (display/etb/booster) for full-size sets (card_count >= 70).
  load   : read the Excel back and upsert the `catalog` table. Unticked SKUs are
           pruned (kept manual control). Idempotent.

Run:
  python -m scraper.games.pokemon_catalog export
  python -m scraper.games.pokemon_catalog load
"""
import argparse
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
from ..config import DATA_DIR, LANGUAGES
from ..db import connect, init_db
from . import cardmarket

CATALOG_XLSX = DATA_DIR / "pokemon_catalog.xlsx"
GAME = "pokemon"
# Western languages share TCGdex set codes with Cardmarket's (EN-led) catalog.
WESTERN_LANGS = {"en", "fr"}

# Tracked sealed-product kinds, in display order, with the label used in catalog
# display_name. (Accessories are intentionally excluded — not a product.)
KIND_LABELS = {
    "display":  "Display (Booster Box)",
    "etb":      "Coffret Dresseur d'Élite (ETB)",
    "upc":      "Ultra Premium Collection (UPC)",
    "bundle":   "Bundle / Lot",
    "tripack":  "Tripack (3 boosters)",
    "duopack":  "Duopack (2 boosters)",
    "blister":  "Blister",
    "coffret":  "Coffret / Tin",
    "minitin":  "Mini Tin",
    "booster":  "Booster (sachet)",
}
KINDS = list(KIND_LABELS)
BASELINE_KINDS = ["display", "etb", "booster"]   # pre-ticked for full-size sets
BASELINE_MIN_CARDS = 70

META_COLS = ["set_code", "name", "series", "release_date", "card_count"]
TICKS = {"x", "X", "1", "oui", "yes", "true", "vrai", "✓"}


def _is_ticked(v) -> bool:
    if v is None or (isinstance(v, float) and pd.isna(v)):
        return False
    if isinstance(v, bool):
        return v
    return str(v).strip().lower() in {t.lower() for t in TICKS}


# --------------------------------------------------------------------------- #
# EXPORT
# --------------------------------------------------------------------------- #
def export():
    # Exclude TCG Pocket (series 'tcgp') — digital app sets with no sealed products.
    with connect() as conn:
        rows = conn.execute("""
            SELECT language, set_code, name, series, release_date, card_count
            FROM sets WHERE game = ? AND (series IS NULL OR series != 'tcgp')
            ORDER BY language, release_date, set_code
        """, (GAME,)).fetchall()
    if not rows:
        raise SystemExit("No pokemon sets in DB. Run build_pokemon_sets first.")

    # Existing catalog ticks (so re-export preserves prior manual edits).
    with connect() as conn:
        existing = {(r["language"], r["set_code"], r["kind"])
                    for r in conn.execute(
                        "SELECT language, set_code, kind FROM catalog WHERE game = ?", (GAME,))}
    has_catalog = bool(existing)

    # Cardmarket product-type existence per (Western) set — drives the baseline
    # for en/fr where set codes are shared. JA/KO/ZH fall back to card_count.
    cm_kinds = cardmarket.set_kinds() if cardmarket.PRODUCTS_FILE.exists() else {}

    DATA_DIR.mkdir(parents=True, exist_ok=True)
    with pd.ExcelWriter(CATALOG_XLSX, engine="openpyxl") as xw:
        _instructions_sheet(xw)
        for lang in LANGUAGES:
            lrows = [r for r in rows if r["language"] == lang]
            if not lrows:
                continue
            data = []
            for r in lrows:
                rec = {c: r[c] for c in META_COLS}
                cm = cm_kinds.get(r["set_code"]) if lang in WESTERN_LANGS else None
                cc = r["card_count"] or 0
                for k in KINDS:
                    if has_catalog:
                        mark = "x" if (lang, r["set_code"], k) in existing else ""
                    elif cm is not None:           # Cardmarket-confirmed product types
                        mark = "x" if k in cm else ""
                    else:                          # baseline for non-Western / unmatched sets
                        mark = "x" if (k in BASELINE_KINDS and cc >= BASELINE_MIN_CARDS) else ""
                    rec[k] = mark
                rec["notes"] = ""
                data.append(rec)
            df = pd.DataFrame(data, columns=META_COLS + KINDS + ["notes"])
            df.to_excel(xw, sheet_name=lang, index=False)

    _format(CATALOG_XLSX)
    print(f"Wrote {CATALOG_XLSX}")
    print("Edit the kind columns (x = track), then: python -m scraper.games.pokemon_catalog load")


def _instructions_sheet(xw):
    lines = [
        ["Pokemon catalog — how to edit"],
        [""],
        ["One sheet per language (fr/en/ja/ko/zh). One row per set."],
        ["Put 'x' (or 1) in a product-kind column to TRACK that SKU; clear it to drop it."],
        ["Then run:  python -m scraper.games.pokemon_catalog load"],
        [""],
        ["Kind columns:"],
        *[[f"  {k}", lbl] for k, lbl in KIND_LABELS.items()],
        [""],
        ["Pre-fill (first export only):"],
        ["  en/fr  = actual product types found on Cardmarket for that set."],
        [f"  ja/ko/zh = baseline ({', '.join(BASELINE_KINDS)}) for sets with "
         f"card_count >= {BASELINE_MIN_CARDS} (Cardmarket doesn't cover these)."],
        ["TCG Pocket (digital) sets are excluded — no sealed products."],
        ["Re-exporting preserves whatever is currently in the catalog table."],
    ]
    pd.DataFrame(lines).to_excel(xw, sheet_name="_instructions", index=False, header=False)


def _format(path):
    wb = load_workbook(path)
    bold = Font(bold=True)
    head_fill = PatternFill("solid", fgColor="D9E1F2")
    for name in wb.sheetnames:
        ws = wb[name]
        for c in ws[1]:
            c.font = bold
            c.alignment = Alignment(horizontal="left")
            if name != "_instructions":
                c.fill = head_fill
        if name == "_instructions":
            ws.column_dimensions["A"].width = 60
            ws.column_dimensions["B"].width = 40
            continue
        ws.freeze_panes = "B2"
        widths = {"set_code": 12, "name": 38, "series": 8, "release_date": 13,
                  "card_count": 11, "notes": 30}
        for i, col in enumerate(ws[1], start=1):
            ws.column_dimensions[get_column_letter(i)].width = widths.get(col.value, 9)
    wb.save(path)


# --------------------------------------------------------------------------- #
# LOAD
# --------------------------------------------------------------------------- #
def load():
    if not CATALOG_XLSX.exists():
        raise SystemExit(f"{CATALOG_XLSX} not found. Run `export` first.")
    init_db()
    xls = pd.read_excel(CATALOG_XLSX, sheet_name=None)
    # set name lookup for display_name
    with connect() as conn:
        names = {(r["language"], r["set_code"]): r["name"] for r in conn.execute(
            "SELECT language, set_code, name FROM sets WHERE game = ?", (GAME,))}

    wanted = []  # (language, set_code, kind, display_name)
    for sheet, df in xls.items():
        if sheet not in LANGUAGES or df.empty:
            continue
        for _, r in df.iterrows():
            code = str(r.get("set_code") or "").strip()
            if not code:
                continue
            for k in KINDS:
                if k in df.columns and _is_ticked(r.get(k)):
                    nm = names.get((sheet, code)) or code
                    wanted.append((sheet, code, k, f"{nm} — {KIND_LABELS[k]} [{sheet.upper()}]"))

    with connect() as conn:
        conn.executemany("""
            INSERT INTO catalog (game, language, set_code, kind, display_name)
            VALUES (?, ?, ?, ?, ?)
            ON CONFLICT(game, language, set_code, kind)
            DO UPDATE SET display_name = excluded.display_name
        """, [(GAME, l, c, k, dn) for (l, c, k, dn) in wanted])
        # Prune catalog rows no longer ticked.
        keep = {(l, c, k) for (l, c, k, _) in wanted}
        existing = conn.execute(
            "SELECT id, language, set_code, kind FROM catalog WHERE game = ?", (GAME,)).fetchall()
        pruned = 0
        for row in existing:
            if (row["language"], row["set_code"], row["kind"]) not in keep:
                conn.execute("DELETE FROM catalog WHERE id = ?", (row["id"],))
                pruned += 1
        conn.commit()
        total = conn.execute("SELECT COUNT(*) FROM catalog WHERE game = ?", (GAME,)).fetchone()[0]
        by_lang = conn.execute("""
            SELECT language, COUNT(*) FROM catalog WHERE game = ?
            GROUP BY language ORDER BY language
        """, (GAME,)).fetchall()
    print(f"Catalog upserted: {len(wanted)} ticked, pruned {pruned}. Total pokemon SKUs: {total}")
    for lang, c in by_lang:
        print(f"  {lang:<4} {c}")


if __name__ == "__main__":
    ap = argparse.ArgumentParser()
    ap.add_argument("cmd", choices=["export", "load"])
    args = ap.parse_args()
    {"export": export, "load": load}[args.cmd]()
