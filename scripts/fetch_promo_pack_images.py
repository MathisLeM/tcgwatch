"""Attach a sealed-pack product image to every entry of optcg_promo_packs.json.

Cardmarket blocks image hotlinking (403); Limitless has no sealed-pack shot.
TCGplayer does, on a fully predictable CDN, and tcgcsv.com mirrors the TCGplayer
catalogue (productId + imageUrl) as free static JSON.

Pipeline
--------
1. Download (cached) the TCGplayer "One Piece Promotion Cards" group (id 17675,
   category 68) from tcgcsv.com -> sealed-pack name -> productId index.
2. For each pack in data/optcg_promo_packs.json, normalised-match (preferring the
   Cardmarket product name, which mirrors TCGplayer's naming) and attach
   `tcgplayer_id` + `image_url` (1000x1000).
3. Download images to images/promo_packs/<tcgplayer_id>.jpg.
4. Rewrite the JSON in place.

Run:  python scripts/fetch_promo_pack_images.py [--no-download]
"""
from __future__ import annotations
import json, re, os, time, argparse, urllib.request

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
JSON_PATH = os.path.join(ROOT, "data", "optcg_promo_packs.json")
GROUPS_CACHE = os.path.join(ROOT, "data", "reference", "tcgplayer_promo_products.json")
IMG_DIR = os.path.join(ROOT, "images", "promo_packs")
UA = "Mozilla/5.0 (Windows NT 10.0; Win64; x64)"
TCGCSV_GROUP = "https://tcgcsv.com/tcgplayer/68/17675/products"   # One Piece -> Promotion Cards
# A handful of sealed packs live in set-specific Release-Event groups, not 17675.
EXTRA_GROUPS = ["24677", "24638", "24579", "24406", "24242", "24068", "23304"]
CDN = "https://tcgplayer-cdn.tcgplayer.com/product/{}_in_1000x1000.jpg"

CODE_PAREN = re.compile(r"\([A-Z0-9]+-\d+\)")

# Packs whose official name differs too much for fuzzy matching: pack_name -> productId.
NAME_TO_PID = {
    "3rd Anniversary! Treasure Campaign Pack": 661447,
}

def norm(s: str) -> str:
    s = (s or "").lower().replace("vol.", "vol").replace("vol ", "vol")
    s = re.sub(r"[^a-z0-9]+", " ", s)
    s = s.replace("championship", "cs")
    return " ".join(s.split())

def fetch_json(url, fp):
    if os.path.exists(fp) and os.path.getsize(fp) > 500:
        return json.load(open(fp, encoding="utf-8"))
    req = urllib.request.Request(url, headers={"User-Agent": UA})
    data = json.loads(urllib.request.urlopen(req, timeout=30).read().decode("utf-8"))
    json.dump(data, open(fp, "w", encoding="utf-8"), ensure_ascii=False)
    time.sleep(0.5)
    return data

def build_index():
    """norm(name) -> productId for sealed products (name without a card code)."""
    idx = {}
    results = fetch_json(TCGCSV_GROUP, GROUPS_CACHE).get("results", [])
    for g in EXTRA_GROUPS:
        fp = os.path.join(ROOT, "data", "reference", f"tcgplayer_group_{g}.json")
        try:
            results += fetch_json(f"https://tcgcsv.com/tcgplayer/68/{g}/products", fp).get("results", [])
        except Exception:
            pass
    for p in results:
        name = p["name"]
        if CODE_PAREN.search(name):      # skip singles
            continue
        idx.setdefault(norm(name), p["productId"])
    return idx

def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--no-download", action="store_true", help="match + write JSON, skip image download")
    args = ap.parse_args()

    idx = build_index()
    print(f"TCGplayer sealed-product index: {len(idx)} noms")

    doc = json.load(open(JSON_PATH, encoding="utf-8"))
    os.makedirs(IMG_DIR, exist_ok=True)

    matched = downloaded = 0
    unmatched = []
    for p in doc["packs"]:
        pid = NAME_TO_PID.get(p["pack_name"])
        keys = [norm(p.get("cm_name") or ""), norm(p["pack_name"])]
        if not pid:
            pid = next((idx[k] for k in keys if k in idx), None)
        if not pid:   # relaxed: drop online/offline qualifier
            for k in keys:
                rk = norm(k.replace("online", "").replace("offline", ""))
                relaxed = {norm(n.replace("online", "").replace("offline", "")): v for n, v in idx.items()}
                if rk in relaxed:
                    pid = relaxed[rk]; break
        if not pid:
            p["tcgplayer_id"] = None
            p["image_url"] = None
            p["image_file"] = None
            unmatched.append(p["pack_name"])
            continue
        matched += 1
        p["tcgplayer_id"] = pid
        p["image_url"] = CDN.format(pid)
        fname = f"{pid}.jpg"
        p["image_file"] = f"images/promo_packs/{fname}"
        dest = os.path.join(IMG_DIR, fname)
        if not args.no_download and not os.path.exists(dest):
            try:
                req = urllib.request.Request(CDN.format(pid), headers={"User-Agent": UA, "Accept": "image/*"})
                blob = urllib.request.urlopen(req, timeout=25).read()
                if blob:
                    open(dest, "wb").write(blob)
                    downloaded += 1
                time.sleep(0.4)
            except Exception as exc:
                print("  ERR img", pid, exc)

    json.dump(doc, open(JSON_PATH, "w", encoding="utf-8"), ensure_ascii=False, indent=2)
    print(f"packs appARies a une image: {matched}/{len(doc['packs'])} | images telechargees: {downloaded}")
    if not args.no_download:
        write_excel(doc)
    if unmatched:
        print(f"sans image ({len(unmatched)}):")
        for n in unmatched:
            print("   -", n)


# ----------------------------------------------------------------- Excel + thumbnails
XLSX_PATH = os.path.join(ROOT, "data", "optcg_promo_packs.xlsx")
THUMB_DIR = os.path.join(IMG_DIR, "thumbs")
THUMB_W = 70   # px

def _thumb(pid):
    """Make/return a small thumbnail of the pack image for embedding."""
    from PIL import Image as PILImage
    src = os.path.join(IMG_DIR, f"{pid}.jpg")
    if not os.path.exists(src):
        return None
    os.makedirs(THUMB_DIR, exist_ok=True)
    dst = os.path.join(THUMB_DIR, f"{pid}.png")
    if not os.path.exists(dst):
        im = PILImage.open(src).convert("RGB")
        h = round(im.height * THUMB_W / im.width)
        im.resize((THUMB_W, h)).save(dst, "PNG")
    return dst

def write_excel(doc):
    import openpyxl
    from openpyxl.styles import Font, Alignment
    from openpyxl.drawing.image import Image as XLImage

    wb = openpyxl.Workbook()
    ws1 = wb.active
    ws1.title = "Packs"
    ws1.append(["Image", "Pack", "Source", "Nb cartes", "Cartes listees", "CM id", "CM nom",
                "Prix trend (EUR)", "Nb variantes", "TCGplayer id", "Limitless URL"])
    for p in doc["packs"]:
        ws1.append(["", p["pack_name"], p["source"], p["card_count"],
                    "oui" if p["cards_listed"] else "non", p["cm_id"], p["cm_name"],
                    p["cm_price_trend"], len(p["cm_variants"]), p.get("tcgplayer_id"),
                    p["limitless_url"]])
        r = ws1.max_row
        thumb = _thumb(p["tcgplayer_id"]) if p.get("tcgplayer_id") else None
        if thumb:
            img = XLImage(thumb)
            ws1.row_dimensions[r].height = max(img.height, 50) * 0.78  # px -> pt approx
            ws1.add_image(img, f"A{r}")
    ws1.column_dimensions["A"].width = THUMB_W / 7.0

    ws2 = wb.create_sheet("Cartes")
    ws2.append(["Pack", "Code", "Nom", "Version (Limitless)", "CM single ids (Unnumbered Promos)",
                "Edite dans", "Note", "Limitless URL"])
    for p in doc["packs"]:
        for c in p["cards"]:
            ws2.append([p["pack_name"], c["code"], c["name"], c["limitless_version"],
                        ", ".join(map(str, c["cm_single_ids"])), c["cm_expansion"],
                        c["note"], c["limitless_url"]])

    for ws in (ws1, ws2):
        for cell in ws[1]:
            cell.font = Font(bold=True)
        ws.freeze_panes = "A2"
        for col in ws.columns:
            letter = col[0].column_letter
            if letter == "A" and ws is ws1:
                continue
            w = max((len(str(c.value)) for c in col if c.value is not None), default=10)
            ws.column_dimensions[letter].width = min(w + 2, 60)
    ws1["A1"].alignment = Alignment(horizontal="center")
    wb.save(XLSX_PATH)
    print(f"OK -> {XLSX_PATH} (vignettes incrustees dans l'onglet Packs)")

if __name__ == "__main__":
    main()
