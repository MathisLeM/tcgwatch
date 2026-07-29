"""Build the OPTCG promotional sealed-pack -> cards listing.

For every promotional sealed pack (Welcome / Tournament / Event / Promotion /
Regional / Dash / Battle / Store-Championship / Championship pack ...) it lists
the promo cards inside, in FR-when-available naming, with Cardmarket product +
price and a Limitless link per card.

Sources
-------
- Contents : onepiece.limitlesstcg.com  (one page per promo pack; scraped + cached)
- Products / prices / FR names : Cardmarket dump in data/<DUMP>/

Key rules (validated with the user)
-----------------------------------
- Card disambiguation: keep ONLY the promo printing => the Cardmarket single
  whose idExpansion == 5303 ("Unnumbered Promos"), never the base-set printing.
- One row per pack, language preference FR > EN > JA among Cardmarket variants.
- Cards are listed only when a pack has <= 20 distinct cards (skip huge promo
  sets). Cardmarket promo products that have no Limitless contents page are
  appended as packs-only (no card list).

Run
---
    python scripts/build_promo_packs.py
Outputs: data/optcg_promo_packs.json  and  data/optcg_promo_packs.xlsx
"""
from __future__ import annotations
import json, re, os, time, urllib.request
from collections import defaultdict

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
DUMP = os.path.join(ROOT, "data", "OP26062026")
CACHE = os.path.join(ROOT, "data", "reference", "limitless_promo_cache")
OUT_JSON = os.path.join(ROOT, "data", "optcg_promo_packs.json")
OUT_XLSX = os.path.join(ROOT, "data", "optcg_promo_packs.xlsx")
os.makedirs(CACHE, exist_ok=True)

UA = "Mozilla/5.0 (Windows NT 10.0; Win64; x64)"
LIM = "https://onepiece.limitlesstcg.com/cards/"
UNNUMBERED_PROMOS_EXP = 5303
MAX_CARDS = 20

# Random-insert / participation promo packs that have a Limitless contents page.
SLUGS = [
    "welcome-pack-02", "welcome-pack-2026-1",
    "tournament-pack-01", "tournament-pack-02", "tournament-pack-03", "tournament-pack-04",
    "tournament-pack-05", "tournament-pack-06", "tournament-pack-07", "tournament-pack-08",
    "tournament-pack-09", "tournament-pack-10", "tournament-pack-11", "tournament-pack-12",
    "tournament-pack-13", "tournament-pack-14",
    "event-pack-01", "event-pack-02", "event-pack-03", "event-pack-04", "event-pack-05",
    "event-pack-06", "event-pack-07", "event-pack-09",
    "promotion-pack-01", "promotion-pack-02", "promotion-card-set-2025",
    "regional-participation-pack-01", "regional-participation-pack-02",
    "regional-participation-pack-2024-1", "regional-participation-pack-2024-2",
    "regional-participation-pack-2024-3", "regional-participation-pack-2025-1",
    "regional-participation-pack-2025-2", "regional-participation-pack-2026-1",
    "store-championship-participation-pack-01", "store-championship-participation-pack-02",
    "dash-pack-01", "dash-pack-eb03", "dash-pack-op14",
    "sealed-battle-kit-01", "uta-deck-battle-participation-pack",
    "championship-2023-celebration-pack", "championship-2023-event-pack",
    "championship-2024-celebration-pack", "championship-2024-event-pack",
    "championship-25-26-celebration-pack", "championship-25-26-event-pack",
    "3rd-anniversary-treasure-campaign-pack", "film-red-promotion-card-set",
]

# Curated aliases: Limitless slug -> Cardmarket base product name (fuzzy match fails).
SLUG_ALIASES = {
    "store-championship-participation-pack-01": "Store Championship 2023 Participation Pack Vol.1",
    "store-championship-participation-pack-02": "Store Championship 2023 Participation Pack Vol.2",
}

# --------------------------------------------------------------------- fetch
def _get(url, fp):
    if os.path.exists(fp) and os.path.getsize(fp) > 500:
        return open(fp, encoding="utf-8").read()
    req = urllib.request.Request(url, headers={"User-Agent": UA})
    html = urllib.request.urlopen(req, timeout=25).read().decode("utf-8", "replace")
    open(fp, "w", encoding="utf-8").write(html)
    time.sleep(0.8)
    return html

CODE_RE = re.compile(r'/cards/([A-Z][A-Z0-9]*-\d+)(?:\?v=(\d+))?')
TITLE_RE = re.compile(r'<title>(.*?)</title>', re.S)

def scrape_pack(slug):
    html = _get(LIM + slug, os.path.join(CACHE, slug + ".html"))
    seen = {}
    for code, v in CODE_RE.findall(html):
        seen.setdefault(code, set()).add(int(v) if v else 0)
    cards = [{"code": c, "v": v} for c in sorted(seen) for v in sorted(seen[c])]
    t = TITLE_RE.search(html)
    title = re.sub(r"\s+", " ", t.group(1)).strip() if t else slug
    title = title.replace(" – Limitless One Piece", "").replace("&#039;", "'")
    return {"slug": slug, "title": title, "n": len(cards), "cards": cards}

NAME_FROM_TITLE = re.compile(r'^(.*?)\s*\(')
def limitless_card_name(code):
    try:
        html = _get(LIM + code, os.path.join(CACHE, "card_" + code + ".html"))
    except Exception:
        return None
    t = TITLE_RE.search(html)
    if not t:
        return None
    m = NAME_FROM_TITLE.match(t.group(1).strip())
    return m.group(1).strip() if m else None

# ----------------------------------------------------------------- load CM dump
def load(name):
    with open(os.path.join(DUMP, name), encoding="utf-8") as f:
        return json.load(f)

ns = load("products_nonsingles_op_2606.json")["products"]
sg = load("products_singles_op_2606.json")["products"]
pg = {p["idProduct"]: p for p in load("price_guide_op_2606.json")["priceGuides"]}

def price(pid):
    g = pg.get(pid)
    return (g.get("trend") or g.get("avg") or g.get("avg7")) if g else None

CODE_IN_NAME = re.compile(r"\(([A-Z0-9]+-\d+)\)")
promo_singles = defaultdict(list)   # code -> [{cm_single_id, name}]  (exp 5303 only)
name_any = {}
for p in sg:
    m = CODE_IN_NAME.search(p["name"])
    if not m:
        continue
    code = m.group(1)
    nm = CODE_IN_NAME.sub("", p["name"]).strip()
    name_any.setdefault(code, nm)
    if p["idExpansion"] == UNNUMBERED_PROMOS_EXP:
        promo_singles[code].append({"cm_single_id": p["idProduct"], "name": nm})

# --------------------------------------------------- CM promo-pack universe / variants
PACK_KW = ["tournament pack", "welcome pack", "event pack", "promotion pack", "promotion card set",
           "regional participation", "store championship", "dash pack", "sealed battle kit",
           "battle participation pack", "celebration pack", "event pack finalist", "top players pack",
           "top player pack", "film red promotion", "release event pack", "anniversary tournament pack",
           "treasure campaign", "participation pack"]
def is_pack(name):
    n = name.lower()
    if "winner" in n or "demo deck" in n or "deck pack" in n:
        return False
    return any(k in n for k in PACK_KW)

LANG_TAGS = re.compile(r"\s*\((Non-English|Asia Region Legal|French Version|English Version|August Pre-Order)\)", re.I)
def base_name(name):
    return LANG_TAGS.sub("", name).strip()
def lang_rank(name):
    n = name.lower()
    if "french version" in n:
        return 0                                      # FR
    if "(non-english)" not in n and "asia region legal" not in n:
        return 1                                      # EN / international
    return 2                                          # JA / Asia

groups = defaultdict(list)
for p in ns:
    if is_pack(p["name"]):
        groups[base_name(p["name"])].append(p)

def norm(s):
    s = s.lower().replace("vol.", "vol").replace("vol ", "vol")
    s = re.sub(r"[^a-z0-9]+", " ", s)
    s = s.replace("championship", "cs").replace("participation", "part").replace("online", "").replace("offline", "")
    return " ".join(s.split())

def find_cm_group(title):
    tn = norm(title)
    best = None
    for base, prods in groups.items():
        bn = norm(base)
        if bn == tn:
            return base, prods
        if tn and (tn in bn or bn in tn) and (best is None or len(bn) < len(norm(best[0]))):
            best = (base, prods)
    return best if best else (None, None)

def pick_variant(prods):
    pref = sorted(prods, key=lambda p: (lang_rank(p["name"]), p["idProduct"]))[0]
    variants = [{"cm_id": p["idProduct"], "cm_name": p["name"], "price": price(p["idProduct"])}
                for p in sorted(prods, key=lambda p: lang_rank(p["name"]))]
    return pref, variants

# ------------------------------------------------------------------- build packs
print("Scraping Limitless (cached):", len(SLUGS), "packs...")
lim = [scrape_pack(s) for s in SLUGS]

out_packs, matched = [], set()
for r in lim:
    if r["slug"] in SLUG_ALIASES and SLUG_ALIASES[r["slug"]] in groups:
        base, prods = SLUG_ALIASES[r["slug"]], groups[SLUG_ALIASES[r["slug"]]]
    else:
        base, prods = find_cm_group(r["title"])
    cm_id = cm_name = cm_price = None
    variants = []
    if prods:
        pref, variants = pick_variant(prods)
        cm_id, cm_name, cm_price = pref["idProduct"], pref["name"], price(pref["idProduct"])
        matched.update(v["cm_id"] for v in variants)

    cards = []
    for c in r["cards"]:
        code, v = c["code"], c["v"]
        cand = promo_singles.get(code, [])
        nm = (cand[0]["name"] if cand else name_any.get(code)) or limitless_card_name(code)
        e = {"code": code, "name": nm, "limitless_version": v,
             "limitless_url": LIM + code + (f"?v={v}" if v else ""),
             "cm_single_ids": [x["cm_single_id"] for x in cand],
             "cm_expansion": "Unnumbered Promos" if cand else None, "note": ""}
        if len(cand) > 1:
            e["note"] = "plusieurs versions promo dans Unnumbered Promos - lien CM a confirmer"
        elif not cand:
            e["note"] = "pas de single promo dans le dump (code P-### ou hors Unnumbered Promos)"
        cards.append(e)

    out_packs.append({
        "pack_name": r["title"], "source": "limitless",
        "limitless_slug": r["slug"], "limitless_url": LIM + r["slug"],
        "cm_id": cm_id, "cm_name": cm_name, "cm_price_trend": cm_price,
        "cm_variants": variants,
        "card_count": r["n"], "cards_listed": r["n"] <= MAX_CARDS,
        "cards": cards if r["n"] <= MAX_CARDS else [],
    })

# packs-only: CM promo products without a Limitless contents page
for base, prods in groups.items():
    if any(p["idProduct"] in matched for p in prods):
        continue
    pref, variants = pick_variant(prods)
    out_packs.append({
        "pack_name": base, "source": "cardmarket-only",
        "limitless_slug": None, "limitless_url": None,
        "cm_id": pref["idProduct"], "cm_name": pref["name"],
        "cm_price_trend": price(pref["idProduct"]), "cm_variants": variants,
        "card_count": None, "cards_listed": False, "cards": [],
    })

doc = {
    "generated_at": time.strftime("%Y-%m-%d %H:%M:%S"),
    "game": "One Piece Card Game",
    "rules": {"promo_expansion": "Unnumbered Promos", "promo_expansion_id": UNNUMBERED_PROMOS_EXP,
              "max_cards_listed": MAX_CARDS, "language_preference": "FR > EN > JA"},
    "sources": {"contents": "onepiece.limitlesstcg.com", "products_prices": os.path.basename(DUMP)},
    "pack_count": len(out_packs), "packs": out_packs,
}
with open(OUT_JSON, "w", encoding="utf-8") as f:
    json.dump(doc, f, ensure_ascii=False, indent=2)

# ------------------------------------------------------------------- Excel
import openpyxl
from openpyxl.styles import Font
wb = openpyxl.Workbook()
ws1 = wb.active
ws1.title = "Packs"
ws1.append(["Pack", "Source", "Nb cartes", "Cartes listees", "CM id", "CM nom",
            "Prix trend (EUR)", "Nb variantes", "Limitless URL"])
for p in out_packs:
    ws1.append([p["pack_name"], p["source"], p["card_count"], "oui" if p["cards_listed"] else "non",
                p["cm_id"], p["cm_name"], p["cm_price_trend"], len(p["cm_variants"]), p["limitless_url"]])

ws2 = wb.create_sheet("Cartes")
ws2.append(["Pack", "Code", "Nom", "Version (Limitless)", "CM single ids (Unnumbered Promos)",
            "Edite dans", "Note", "Limitless URL"])
for p in out_packs:
    for c in p["cards"]:
        ws2.append([p["pack_name"], c["code"], c["name"], c["limitless_version"],
                    ", ".join(map(str, c["cm_single_ids"])), c["cm_expansion"], c["note"], c["limitless_url"]])

for ws in (ws1, ws2):
    for cell in ws[1]:
        cell.font = Font(bold=True)
    ws.freeze_panes = "A2"
    for col in ws.columns:
        w = max((len(str(c.value)) for c in col if c.value is not None), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(w + 2, 60)
wb.save(OUT_XLSX)

n_cards = sum(1 for p in out_packs if p["cards"])
print(f"OK -> {OUT_JSON}")
print(f"OK -> {OUT_XLSX}")
print(f"{len(out_packs)} packs ({n_cards} avec cartes detaillees, "
      f"{len(out_packs)-n_cards} packs-only) | "
      f"{sum(len(p['cards']) for p in out_packs)} lignes-cartes")
